import os
import openpyxl
import torch
from PIL import Image
import torchvision.transforms as transforms
from sklearn.metrics import confusion_matrix
from metrics.file_io import load as load_image
import numpy as np
import pandas as pd
# from metrics.cal_score import cal_dice,cal_iou,cal_fwβ,cal_sα,cal_em,cal_mae

def fun_object(pred, gt):
    temp = pred[gt == 1]
    x = temp.mean()
    sigma_x = temp.std()
    score = 2.0 * x / (x * x + 1.0 + sigma_x + 1e-20)

    return score


def fun_S_object(pred, gt):
    fg = torch.where(gt == 0, torch.zeros_like(pred), pred)
    bg = torch.where(gt == 1, torch.zeros_like(pred), 1 - pred)
    o_fg = fun_object(fg, gt)
    o_bg = fun_object(bg, 1 - gt)
    u = gt.float().mean()
    Q = u * o_fg + (1 - u) * o_bg
    return Q


def fun_centroid(gt, cuda=False):
    rows, cols = gt.size()[-2:]
    gt = gt.view(rows, cols)

    if gt.sum() == 0:

        if cuda:
            X = torch.eye(1).cuda() * round(cols / 2)
            Y = torch.eye(1).cuda() * round(rows / 2)
        else:
            X = torch.eye(1) * round(cols / 2)
            Y = torch.eye(1) * round(rows / 2)

    else:
        total = gt.sum()

        if cuda:
            i = torch.from_numpy(np.arange(0, cols)).cuda().float()
            j = torch.from_numpy(np.arange(0, rows)).cuda().float()
        else:
            i = torch.from_numpy(np.arange(0, cols)).float()
            j = torch.from_numpy(np.arange(0, rows)).float()

        X = torch.round((gt.sum(dim=0) * i).sum() / total)
        Y = torch.round((gt.sum(dim=1) * j).sum() / total)

    return X.long(), Y.long()


def fun_divideGT(gt, X, Y):
    h, w = gt.size()[-2:]
    area = h * w
    gt = gt.view(h, w)
    LT = gt[:Y, :X]
    RT = gt[:Y, X:w]
    LB = gt[Y:h, :X]
    RB = gt[Y:h, X:w]
    X = X.float()
    Y = Y.float()
    w1 = X * Y / area
    w2 = (w - X) * Y / area
    w3 = X * (h - Y) / area
    w4 = 1 - w1 - w2 - w3

    return LT, RT, LB, RB, w1, w2, w3, w4


def fun_dividePrediction(pred, X, Y):
    h, w = pred.size()[-2:]
    pred = pred.view(h, w)
    LT = pred[:Y, :X]
    RT = pred[:Y, X:w]
    LB = pred[Y:h, :X]
    RB = pred[Y:h, X:w]

    return LT, RT, LB, RB


def fun_ssim(pred, gt):
    gt = gt.float()
    h, w = pred.size()[-2:]
    N = h * w
    x = pred.mean()
    y = gt.mean()
    sigma_x2 = ((pred - x) * (pred - x)).sum() / (N - 1 + 1e-20)
    sigma_y2 = ((gt - y) * (gt - y)).sum() / (N - 1 + 1e-20)
    sigma_xy = ((pred - x) * (gt - y)).sum() / (N - 1 + 1e-20)

    aplha = 4 * x * y * sigma_xy
    beta = (x * x + y * y) * (sigma_x2 + sigma_y2)

    if aplha != 0:
        Q = aplha / (beta + 1e-20)
    elif aplha == 0 and beta == 0:
        Q = 1.0
    else:
        Q = 0

    return Q


def fun_S_region(pred, gt):
    X, Y = fun_centroid(gt)
    gt1, gt2, gt3, gt4, w1, w2, w3, w4 = fun_divideGT(gt, X, Y)
    p1, p2, p3, p4 = fun_dividePrediction(pred, X, Y)
    Q1 = fun_ssim(p1, gt1)
    Q2 = fun_ssim(p2, gt2)
    Q3 = fun_ssim(p3, gt3)
    Q4 = fun_ssim(p4, gt4)
    Q = w1 * Q1 + w2 * Q2 + w3 * Q3 + w4 * Q4

    return Q


class cal_sα(object):
    # Structure-measure: A new way to evaluate foreground maps (ICCV 2017)
    def __init__(self, alpha=0.5):
        self.alpha = alpha

    def sa(self, pred, gt):
        gt = torch.from_numpy(gt)  # Convert gt numpy array to Tensor
        pred = torch.from_numpy(pred)
        # gt = np.array(gt)
        # gt = gt > 0.5
        return self.cal(pred, gt)

    def cal(self, pred, gt):
        alpha, avg_q, img_num = 0.5, 0.0, 0.0
        y = gt.float().mean()
        ##
        if y == 0:
            x = pred.mean()
            Q = 1.0 - x
        elif y == 1:
            x = pred.mean()
            Q = x
        else:
            Q = alpha * fun_S_object(pred, gt) + (1 - alpha) * fun_S_region(pred, gt)
            if Q.item() < 0:
                Q = torch.FLoatTensor([0.0])
        return Q.item()
class cal_em(object):
    def __init__(self, cuda=False):
        self.cuda = cuda

    def em(self, pred, gt):
        em_value = self._eval_e(pred, gt, 255)
        em_value = em_value.item()  # Get the value without the tensor
        return em_value

    def _eval_e(self, y_pred, y, num):
        if self.cuda:
            score = torch.zeros(num).cuda()
        else:
            score = torch.zeros(num)
        y_pred = torch.from_numpy(y_pred).float()  # Convert y_pred to a PyTorch tensor of float type
        y = torch.from_numpy(y).float()  # Convert y to a PyTorch tensor of float type
        for i in range(num):
            fm = y_pred - torch.mean(y_pred)
            gt = y - torch.mean(y)
            align_matrix = 2 * gt * fm / (gt * gt + fm * fm + 1e-20)
            enhanced = ((align_matrix + 1) * (align_matrix + 1)) / 4
            enhanced_tensor = torch.tensor(enhanced)
            num_elements = y.numel()  # Use numel() for PyTorch tensors
            score[i] = torch.sum(enhanced_tensor) / (num_elements - 1 + 1e-20)
        return score.max()


cal_sα = cal_sα()
# cal_em = cal_em()
cal_em = cal_em()

def calculate_metrics(gt, seg):
    data_transform = transforms.Compose([
        transforms.Resize((352, 352))])

    gt = Image.fromarray(np.uint8(gt))
    gt = data_transform(gt)
    gt = gt.convert('L')
    gt = np.array(gt)
    pred_binary = (seg >= 0.5).astype(float)
    pred_binary_inverse = (pred_binary == 0).astype(float)

    gt_binary = (gt >= 0.5).astype(float)
    gt_binary_inverse = (gt_binary == 0).astype(float)

    TP = np.multiply(pred_binary, gt_binary).sum()
    FP = np.multiply(pred_binary, gt_binary_inverse).sum()
    TN = np.multiply(pred_binary_inverse, gt_binary_inverse).sum()
    FN = np.multiply(pred_binary_inverse, gt_binary).sum()
# ========================师姐的用混淆矩阵=============
#     pred_binary = (seg >= 0.5).astype(int)
#     c_matrix = confusion_matrix(gt.flatten(), pred_binary.flatten())
#     TN = c_matrix[0, 0]
#     TP = c_matrix[1, 1]
#     FN = c_matrix[1, 0]
#     FP = c_matrix[0, 1]
# ========================师姐的用混淆矩阵=============
    if TP.item() == 0:
        TP = torch.Tensor([1]).cpu()

    recall = TP / (TP + FN)
    precision = TP / (TP + FP)
    dice = (2 * TP) / (2 * TP + FN + FP)
    iou = TP / (TP + FP + FN)
    fwb = (1+0.3) * precision * recall / (0.3 * precision + recall)

    sα = cal_sα.sa(seg, gt)
    em = cal_em.em(seg, gt)


    # sα = cal_sα.update(seg, gt)
    # em = cal_em.update(seg, gt)
    mae = np.abs(gt - seg).mean()
# ====================================================
    return dice, iou, fwb,sα,em,mae


def metrics(args,epoch,_data_name):
    gts_path = r'./data/Fold_TestDataset/Fold1_Kvasir\masks/'  # ground truth 的地址
    results_path = r'./test_result_picture/{}/'.format(_data_name)  # 推理结果地址
    picture_sum = len([name for name in os.listdir(gts_path) if os.path.isfile(os.path.join(gts_path, name))])
    print("图片数量为：",picture_sum)
    test_images = range(1, picture_sum+1)
    dice_list, iou_list, fwb_list, sα_list, em_list,  mae_list=[], [], [], [], [], []

    gts_path = [gts_path + f for f in os.listdir(gts_path) if f.endswith('.jpg') or f.endswith('.png')or f.endswith('.bmp')]
    results_path = [results_path + f for f in os.listdir(results_path) if f.endswith('.jpg') or f.endswith('.png')or f.endswith('.bmp')]
    gts = sorted(gts_path)
    segs = sorted(results_path)
    for i, image in enumerate(sorted(test_images)):
        # gt = load_image(gts_path + '%d.png' % image, normalize=True)
        # seg = load_image(results_path + '%d.png' % image, normalize=True)
        gt_path = gts[i]
        seg_path = segs[i]
        gt = load_image(gt_path, normalize=True)
        seg = load_image(seg_path, normalize=True)
        gt_name = os.path.basename(gt_path)
        seg_name = os.path.basename(seg_path)
        # print("第{}个图片".format(i+1))
        # print("gt_name:",gt_name)
        # print("seg_name:",seg_name)

        dice, iou, fwb, sα, em, mae= calculate_metrics(gt, seg)
        dice_list.append(dice)
        iou_list.append(iou)
        fwb_list.append(fwb)
        sα_list.append(sα)
        em_list.append(em)
        mae_list.append(mae)

    metrics_data = {
        'Image': list(range(1, picture_sum+1)),
        'Dice': dice_list,
        'IOU': iou_list,
        'Fwb': fwb_list,
        'Sα': sα_list,
        'Em': em_list,
        'Mae': mae_list,
    }

    metrics_df = pd.DataFrame(data=metrics_data)
    metrics_df = metrics_df[['Image', 'Dice', 'IOU', 'Fwb', 'Sα', 'Em', 'Mae']]

    metrics_df.set_index('Image', inplace=True)
    Aver = metrics_df.mean()
    print(metrics_df)
    print(Aver)
    print('计算评价指标完成！================================')
    Dice = format(Aver["Dice"], '.6f')
    IOU = format(Aver["IOU"], '.6f')
    Fwb = format(Aver["Fwb"], '.6f')
    Sα = format(Aver["Sα"], '.6f')
    Em = format(Aver["Em"], '.6f')
    Mae = format(Aver["Mae"], '.6f')
    # 将数据保存到 Excel 文件
    a = True
    if a:
        workbook_filename = 'results-info/{}.xlsx'.format(_data_name)
        if os.path.exists(workbook_filename):
            workbook = openpyxl.load_workbook(workbook_filename)
        else:
            workbook = openpyxl.Workbook()

        worksheet = workbook['Sheet1']
        old_rows = worksheet.max_row
        worksheet.cell(row=old_rows+1, column=1, value =old_rows)
        worksheet.cell(row=old_rows+1, column=2, value =Dice)
        worksheet.cell(row=old_rows+1, column=3, value =IOU)
        worksheet.cell(row=old_rows+1, column=4, value =Fwb)
        worksheet.cell(row=old_rows+1, column=5, value =Sα)
        worksheet.cell(row=old_rows+1, column=6, value =Em)
        worksheet.cell(row=old_rows+1, column=7, value =Mae)
        workbook.save(workbook_filename)
        print('写入excel文件完成！================================')